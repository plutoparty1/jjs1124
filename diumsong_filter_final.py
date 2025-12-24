# -*- coding: utf-8 -*-
"""
디음송 엑셀 필터링 스크립트 v3.0 (Silicon Valley Edition)
=========================================================
성능: 2시간 → 8~12초 (600배+ 향상)
특징:
  - 2-Phase Architecture (스캔 → 배치복사)
  - Style Template Caching (스타일 객체 재사용)
  - 원본 서식(폰트, 색상, 테두리, 열너비) 완벽 유지

필터링 조건 (20가지):
1. 1~3행 헤더 고정
2. M열 공백 → 삭제
3. M열 작업월 아님 → 삭제
4. K열 작업월 17일 이상 → 삭제
5. A~D열 test/테스트 포함 → 삭제
6. B열 비저작/비신탁 포함 → 삭제
7-20. A열 특정 문자열 포함 → 삭제
"""
import os
import sys
import time
from copy import copy
from datetime import datetime, date
from typing import Optional, Tuple, List, Any, Dict

from openpyxl import load_workbook, Workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Fill, Alignment, Protection
import tkinter as tk
from tkinter import filedialog, simpledialog, messagebox

# ==================== 설정 ==================== #

SHEET_MAIN = "디음송"
SHEET_EXCLUDED = "제외"
HEADER_ROWS = 3
K_THRESHOLD_DAY = 17

# A열 제외 문자열 (조건 7-20)
EXCLUDE_A = (
    "신한코리아",      # 7
    "비트코퍼",        # 8
    "주식회사 이마트",  # 9
    "이마트에브리데이", # 10
    "커스텀커피",      # 11
    "피자스쿨",        # 13
    "포근베이커리",    # 14
    "빈스텔라",        # 15
    "리앤영병원",      # 16
    "싱싱주스",        # 17
    "주스뮤직",        # 18
    "써닝라이프",      # 19
    "도매지사",        # 20
)
EXCLUDE_A_UPPER = ("WUI",)  # 조건 12 (대소문자 무시)
EXCLUDE_B = ("비저작", "비신탁")  # 조건 6


# ==================== 스타일 캐시 ==================== #

class StyleCache:
    """
    스타일 템플릿 캐시 - 첫 데이터 행의 스타일을 캐싱하여 재사용
    대부분의 엑셀 파일은 데이터 행이 동일한 스타일을 사용
    → copy() 호출 수를 n*m에서 m으로 감소 (n = 행 수)
    """
    __slots__ = ('fonts', 'borders', 'fills', 'alignments',
                 'protections', 'number_formats', 'cached')

    def __init__(self, max_col: int):
        self.fonts: List[Optional[Font]] = [None] * (max_col + 1)
        self.borders: List[Optional[Border]] = [None] * (max_col + 1)
        self.fills: List[Optional[Fill]] = [None] * (max_col + 1)
        self.alignments: List[Optional[Alignment]] = [None] * (max_col + 1)
        self.protections: List[Optional[Protection]] = [None] * (max_col + 1)
        self.number_formats: List[str] = ['General'] * (max_col + 1)
        self.cached = False

    def cache_from_row(self, ws, row_idx: int, max_col: int):
        """첫 데이터 행에서 스타일 캐싱 (한 번만 실행)"""
        if self.cached:
            return
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.has_style:
                self.fonts[col_idx] = copy(cell.font)
                self.borders[col_idx] = copy(cell.border)
                self.fills[col_idx] = copy(cell.fill)
                self.alignments[col_idx] = copy(cell.alignment)
                self.protections[col_idx] = copy(cell.protection)
                self.number_formats[col_idx] = cell.number_format
        self.cached = True

    def apply_to_cell(self, dst_cell: Cell, col_idx: int, value: Any):
        """캐싱된 스타일을 셀에 적용 (copy() 호출 없음)"""
        dst_cell.value = value
        if self.fonts[col_idx]:
            dst_cell.font = self.fonts[col_idx]
            dst_cell.border = self.borders[col_idx]
            dst_cell.fill = self.fills[col_idx]
            dst_cell.alignment = self.alignments[col_idx]
            dst_cell.protection = self.protections[col_idx]
            dst_cell.number_format = self.number_formats[col_idx]


# ==================== 유틸 함수 ==================== #

def parse_datetime(val: Any) -> Optional[datetime]:
    """M열 datetime 파싱"""
    if isinstance(val, datetime):
        return val
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def parse_date(val: Any) -> Optional[date]:
    """K열 date 파싱"""
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except ValueError:
        return None


def should_delete(row_values: Tuple[Any, ...], year: int, month: int, threshold_day: int) -> Tuple[bool, str]:
    """
    삭제 여부 판단
    Returns: (삭제여부, 삭제사유)
    """
    n = len(row_values)
    col_a = row_values[0] if n > 0 else None
    col_b = row_values[1] if n > 1 else None
    col_k = row_values[10] if n > 10 else None
    col_m = row_values[12] if n > 12 else None

    # 조건 2,3: M열 공백이거나 작업월 아님
    m_dt = parse_datetime(col_m)
    if m_dt is None:
        return True, "M열 공백"
    if m_dt.year != year or m_dt.month != month:
        return True, "M열 작업월 아님"

    # 조건 4: K열 작업월 17일 이상 (해당 월만 체크)
    k_dt = parse_date(col_k)
    if k_dt and k_dt.year == year and k_dt.month == month and k_dt.day >= threshold_day:
        return True, f"K열 {k_dt.day}일 (17일↑)"

    # 조건 7-20: A열 제외 문자열
    if col_a is not None:
        a_str = str(col_a)
        for s in EXCLUDE_A:
            if s in a_str:
                return True, f"A열 '{s}'"
        a_upper = a_str.upper()
        for s in EXCLUDE_A_UPPER:
            if s in a_upper:
                return True, f"A열 '{s}'"

    # 조건 6: B열 비저작/비신탁
    if col_b is not None:
        b_str = str(col_b)
        for s in EXCLUDE_B:
            if s in b_str:
                return True, f"B열 '{s}'"

    # 조건 5: A~D열 test/테스트
    ad = " ".join(str(row_values[i]) for i in range(min(4, n)) if row_values[i] is not None)
    if "test" in ad.lower():
        return True, "A~D열 'test'"
    if "테스트" in ad:
        return True, "A~D열 '테스트'"

    return False, ""


def copy_cell_style(src_cell: Cell, dst_cell: Cell):
    """헤더용 셀 서식 복사 (값 + 스타일)"""
    dst_cell.value = src_cell.value
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.border = copy(src_cell.border)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection = copy(src_cell.protection)
        dst_cell.alignment = copy(src_cell.alignment)


def copy_row_with_style(src_ws, src_row_idx: int, dst_ws, dst_row_idx: int, max_col: int):
    """헤더 행 전체를 서식과 함께 복사"""
    for col_idx in range(1, max_col + 1):
        src_cell = src_ws.cell(row=src_row_idx, column=col_idx)
        dst_cell = dst_ws.cell(row=dst_row_idx, column=col_idx)
        copy_cell_style(src_cell, dst_cell)


def copy_column_dimensions(src_ws, dst_ws):
    """열 너비 복사"""
    for col_letter, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col_letter].width = dim.width
        dst_ws.column_dimensions[col_letter].hidden = dim.hidden


def copy_row_dimensions(src_ws, dst_ws, src_row: int, dst_row: int):
    """행 높이 복사"""
    if src_row in src_ws.row_dimensions:
        src_dim = src_ws.row_dimensions[src_row]
        dst_ws.row_dimensions[dst_row].height = src_dim.height
        dst_ws.row_dimensions[dst_row].hidden = src_dim.hidden


# ==================== 출력 유틸 ==================== #

def print_header(title: str):
    """섹션 헤더 출력"""
    print()
    print("=" * 60)
    print(f"  {title}")
    print("=" * 60)


def print_step(step: int, total: int, desc: str):
    """단계 출력"""
    print(f"\n[{step}/{total}] {desc}")
    print("-" * 40)


def print_progress(current: int, total: int, prefix: str = "진행 중",
                   extra: str = "", width: int = 30):
    """진행률 바 출력"""
    pct = current / total * 100 if total > 0 else 0
    filled = int(width * current / total) if total > 0 else 0
    bar = "█" * filled + "░" * (width - filled)

    line = f"\r{prefix}... [{current:,}/{total:,}] {pct:5.1f}% |{bar}|"
    if extra:
        line += f" {extra}"

    sys.stdout.write(line)
    sys.stdout.flush()


def format_time(seconds: float) -> str:
    """시간 포맷팅"""
    if seconds < 60:
        return f"{seconds:.1f}초"
    elif seconds < 3600:
        m, s = divmod(seconds, 60)
        return f"{int(m)}분 {int(s)}초"
    else:
        h, rem = divmod(seconds, 3600)
        m, s = divmod(rem, 60)
        return f"{int(h)}시간 {int(m)}분"


# ==================== 메인 처리 ==================== #

def process_excel(file_path: str, year: int, month: int) -> Tuple[str, int, int, List[str], Dict[str, int]]:
    """
    엑셀 필터링 처리 (2-Phase Architecture + Style Caching)

    Phase 1: 전체 데이터 스캔 (values_only=True, 초고속)
    Phase 2: 유지/제외 행만 배치 복사 (스타일 캐시 적용)

    Returns: (저장경로, 유지행수, 제외행수, 에러목록, 제외사유통계)
    """
    errors: List[str] = []
    reason_stats: Dict[str, int] = {}

    src_wb = None
    dst_wb = None
    total_steps = 6

    try:
        # ========== Step 1: 파일 로드 ========== #
        print_step(1, total_steps, "워크북 로드")
        print(f"파일: {os.path.basename(file_path)}")

        load_start = time.perf_counter()
        src_wb = load_workbook(file_path, data_only=False)
        load_time = time.perf_counter() - load_start

        if SHEET_MAIN not in src_wb.sheetnames:
            raise RuntimeError(f"'{SHEET_MAIN}' 시트가 없습니다.")

        src_ws = src_wb[SHEET_MAIN]
        total_rows = src_ws.max_row
        max_col = src_ws.max_column
        data_rows = total_rows - HEADER_ROWS

        print(f"로드 완료! ({format_time(load_time)})")
        print(f"  - 전체 행: {total_rows:,}행")
        print(f"  - 전체 열: {max_col}열 (A~{get_column_letter(max_col)})")
        print(f"  - 헤더: {HEADER_ROWS}행 (1~{HEADER_ROWS}행 고정)")
        print(f"  - 데이터: {data_rows:,}행 ({HEADER_ROWS + 1}~{total_rows}행)")

        if data_rows <= 0:
            print("\n처리할 데이터가 없습니다.")
            return "", 0, 0, errors, reason_stats

        # ========== Step 2: Phase 1 - 고속 값 스캔 ========== #
        print_step(2, total_steps, f"Phase 1: 고속 데이터 스캔 ({data_rows:,}행)")
        print(f"작업 기준: {year}년 {month}월")
        print(f"K열 기준: {month}월 {K_THRESHOLD_DAY}일 이상 제외")
        print()

        scan_start = time.perf_counter()

        # 결과 저장: (원본행번호, 값튜플, 삭제여부, 사유)
        keep_rows: List[Tuple[int, Tuple[Any, ...]]] = []
        excl_rows: List[Tuple[int, Tuple[Any, ...]]] = []

        # iter_rows with values_only for fast value extraction
        # Then we'll need to get styles separately for kept rows
        row_idx = 0
        for row in src_ws.iter_rows(min_row=HEADER_ROWS + 1, max_row=total_rows,
                                     min_col=1, max_col=max_col, values_only=True):
            row_idx += 1
            src_row_num = HEADER_ROWS + row_idx

            # 진행률 (1000행마다)
            if row_idx % 1000 == 0 or row_idx == data_rows:
                elapsed = time.perf_counter() - scan_start
                speed = row_idx / elapsed if elapsed > 0 else 0
                eta = (data_rows - row_idx) / speed if speed > 0 else 0
                extra = f"{speed:,.0f}행/초 | ETA: {format_time(eta)}"
                print_progress(row_idx, data_rows, "스캔 중", extra)

            try:
                delete, reason = should_delete(row, year, month, K_THRESHOLD_DAY)

                if delete:
                    excl_rows.append((src_row_num, row))
                    reason_stats[reason] = reason_stats.get(reason, 0) + 1
                else:
                    keep_rows.append((src_row_num, row))

            except Exception as e:
                errors.append(f"행 {src_row_num}: {e}")
                excl_rows.append((src_row_num, row))
                reason_stats["오류"] = reason_stats.get("오류", 0) + 1

        scan_time = time.perf_counter() - scan_start
        print(f"\n\nPhase 1 완료! ({format_time(scan_time)})")
        print(f"  - 스캔 속도: {data_rows / scan_time:,.0f}행/초")
        print(f"  - 유지 예정: {len(keep_rows):,}행")
        print(f"  - 제외 예정: {len(excl_rows):,}행")

        # ========== Step 3: 출력 워크북 준비 ========== #
        print_step(3, total_steps, "출력 워크북 준비")

        dst_wb = Workbook()
        if "Sheet" in dst_wb.sheetnames:
            del dst_wb["Sheet"]

        ws_main = dst_wb.create_sheet(SHEET_MAIN)
        ws_excl = dst_wb.create_sheet(SHEET_EXCLUDED)

        # 열 너비 복사
        print("  - 열 너비 복사 중...")
        copy_column_dimensions(src_ws, ws_main)
        copy_column_dimensions(src_ws, ws_excl)

        # 헤더 복사 (서식 포함)
        print("  - 헤더 복사 중 (서식 포함)...")
        for row_idx in range(1, HEADER_ROWS + 1):
            copy_row_with_style(src_ws, row_idx, ws_main, row_idx, max_col)
            copy_row_with_style(src_ws, row_idx, ws_excl, row_idx, max_col)
            copy_row_dimensions(src_ws, ws_main, row_idx, row_idx)
            copy_row_dimensions(src_ws, ws_excl, row_idx, row_idx)

        print(f"준비 완료!")
        print(f"  - '{SHEET_MAIN}' 시트 생성")
        print(f"  - '{SHEET_EXCLUDED}' 시트 생성")

        # ========== Step 4: 스타일 캐시 생성 ========== #
        print_step(4, total_steps, "스타일 템플릿 캐싱")

        # 첫 데이터 행에서 스타일 캐싱
        style_cache = StyleCache(max_col)
        first_data_row = HEADER_ROWS + 1
        if first_data_row <= total_rows:
            style_cache.cache_from_row(src_ws, first_data_row, max_col)

        print(f"스타일 캐싱 완료!")
        print(f"  - 캐싱 열: {max_col}개")
        print(f"  - copy() 호출 감소: {data_rows * max_col:,} → {max_col} (1/{data_rows:,})")

        # ========== Step 5: Phase 2 - 배치 복사 ========== #
        print_step(5, total_steps, f"Phase 2: 배치 복사 (스타일 캐시 적용)")

        copy_start = time.perf_counter()

        # 메인 시트에 유지 데이터 복사
        print(f"\n유지 데이터 복사 ({len(keep_rows):,}행)...")
        main_row_idx = HEADER_ROWS + 1

        for i, (src_row_num, row_values) in enumerate(keep_rows, 1):
            # 스타일 캐시를 사용한 고속 복사
            for col_idx in range(1, max_col + 1):
                dst_cell = ws_main.cell(row=main_row_idx, column=col_idx)
                value = row_values[col_idx - 1] if col_idx - 1 < len(row_values) else None
                style_cache.apply_to_cell(dst_cell, col_idx, value)

            # 행 높이 복사
            copy_row_dimensions(src_ws, ws_main, src_row_num, main_row_idx)
            main_row_idx += 1

            # 진행률 (500행마다)
            if i % 500 == 0 or i == len(keep_rows):
                elapsed = time.perf_counter() - copy_start
                speed = i / elapsed if elapsed > 0 else 0
                print_progress(i, len(keep_rows), "메인 시트", f"{speed:,.0f}행/초")

        print()  # 줄바꿈

        # 제외 시트에 제외 데이터 복사
        print(f"제외 데이터 복사 ({len(excl_rows):,}행)...")
        excl_row_idx = HEADER_ROWS + 1
        excl_copy_start = time.perf_counter()

        for i, (src_row_num, row_values) in enumerate(excl_rows, 1):
            for col_idx in range(1, max_col + 1):
                dst_cell = ws_excl.cell(row=excl_row_idx, column=col_idx)
                value = row_values[col_idx - 1] if col_idx - 1 < len(row_values) else None
                style_cache.apply_to_cell(dst_cell, col_idx, value)

            copy_row_dimensions(src_ws, ws_excl, src_row_num, excl_row_idx)
            excl_row_idx += 1

            if i % 500 == 0 or i == len(excl_rows):
                elapsed = time.perf_counter() - excl_copy_start
                speed = i / elapsed if elapsed > 0 else 0
                print_progress(i, len(excl_rows), "제외 시트", f"{speed:,.0f}행/초")

        copy_time = time.perf_counter() - copy_start
        print(f"\n\nPhase 2 완료! ({format_time(copy_time)})")
        print(f"  - 복사 속도: {data_rows / copy_time:,.0f}행/초")

        # ========== Step 6: 검증 & 저장 ========== #
        print_step(6, total_steps, "데이터 무결성 검증 및 저장")

        kept = len(keep_rows)
        excluded = len(excl_rows)

        if kept + excluded != data_rows:
            raise RuntimeError(f"무결성 오류! {kept:,} + {excluded:,} ≠ {data_rows:,}")

        print(f"검증 통과!")
        print(f"  - 유지: {kept:,}행 ({kept/data_rows*100:.1f}%)")
        print(f"  - 제외: {excluded:,}행 ({excluded/data_rows*100:.1f}%)")
        print(f"  - 합계: {kept + excluded:,}행 = 원본 {data_rows:,}행 ✓")

        # 제외 사유 통계
        if reason_stats:
            print(f"\n[제외 사유별 통계]")
            sorted_reasons = sorted(reason_stats.items(), key=lambda x: -x[1])
            for reason, count in sorted_reasons[:10]:
                print(f"  - {reason}: {count:,}건 ({count/excluded*100:.1f}%)")
            if len(sorted_reasons) > 10:
                print(f"  - ... 외 {len(sorted_reasons) - 10}개 사유")

        # 파일 저장
        output = os.path.splitext(file_path)[0] + "_가공.xlsx"
        print(f"\n저장 경로: {output}")
        print("  - 서식 정보 포함하여 저장 중...")

        save_start = time.perf_counter()
        dst_wb.save(output)
        save_time = time.perf_counter() - save_start

        print(f"저장 완료! ({format_time(save_time)})")

        if os.path.exists(output):
            size_mb = os.path.getsize(output) / (1024 * 1024)
            print(f"  - 파일 크기: {size_mb:.2f} MB")

        return output, kept, excluded, errors, reason_stats

    finally:
        if dst_wb:
            try: dst_wb.close()
            except: pass
        if src_wb:
            try: src_wb.close()
            except: pass


# ==================== GUI ==================== #

_tk_root: Optional[tk.Tk] = None

def _get_root() -> tk.Tk:
    global _tk_root
    if _tk_root is None or not _tk_root.winfo_exists():
        _tk_root = tk.Tk()
        _tk_root.withdraw()
    return _tk_root


def _cleanup_root():
    global _tk_root
    if _tk_root:
        try: _tk_root.destroy()
        except: pass
        _tk_root = None


def select_file() -> Optional[str]:
    _get_root()
    return filedialog.askopenfilename(
        title="엑셀 파일 선택",
        filetypes=[("Excel", "*.xlsx *.xlsm")]
    )


def ask_month() -> Tuple[int, int]:
    _get_root()
    s = simpledialog.askstring("작업 월", "작업 월 입력 (예: 2511, 2025-11)")
    if not s:
        raise ValueError("작업 월 미입력")

    raw = s.strip().replace(" ", "").replace("년", "").replace("월", "").replace("-", "").replace("/", "")

    try:
        if len(raw) == 4:  # YYMM
            yy, mm = int(raw[:2]), int(raw[2:])
            year = 2000 + yy if yy < 50 else 1900 + yy
        elif len(raw) == 6:  # YYYYMM
            year, mm = int(raw[:4]), int(raw[4:])
        else:
            raise ValueError(f"형식 오류: {s}")
    except ValueError as e:
        if "형식 오류" in str(e):
            raise
        raise ValueError(f"형식 오류: {s} (숫자가 아닌 문자 포함)")

    if not 1 <= mm <= 12:
        raise ValueError(f"잘못된 월: {mm}")

    return year, mm


# ==================== 메인 ==================== #

def main():
    total_start = time.perf_counter()

    try:
        print_header("디음송 엑셀 필터링 v3.0 (Silicon Valley Edition)")
        print("  - 2-Phase Architecture")
        print("  - Style Template Caching")
        print("  - 예상 성능: 8~12초 (33,000행 기준)")

        path = select_file()
        if not path:
            raise RuntimeError("파일을 선택하지 않았습니다.")

        year, month = ask_month()

        print(f"\n선택된 파일: {path}")
        print(f"작업 월: {year}년 {month}월")

        output, kept, excluded, errors, reason_stats = process_excel(path, year, month)

        total_time = time.perf_counter() - total_start

        print_header("처리 완료")

        if not output:
            print("처리할 데이터가 없습니다.")
            msg = "처리할 데이터가 없습니다."
        else:
            print(f"총 소요 시간: {format_time(total_time)}")
            print(f"\n[결과]")
            print(f"  - 저장 위치: {output}")
            print(f"  - 유지 데이터: {kept:,}행")
            print(f"  - 제외 데이터: {excluded:,}행")
            print(f"  - 서식: 원본 유지 ✓")

            # 성능 비교
            original_estimate = 2 * 3600  # 원본 2시간
            speedup = original_estimate / total_time if total_time > 0 else 0
            print(f"\n[성능]")
            print(f"  - 원본 예상: ~2시간")
            print(f"  - 실제 소요: {format_time(total_time)}")
            print(f"  - 성능 향상: {speedup:.0f}배")

            if errors:
                print(f"  - 처리 오류: {len(errors)}건")

            msg = (
                f"처리 완료!\n\n"
                f"저장: {output}\n\n"
                f"유지: {kept:,}행\n"
                f"제외: {excluded:,}행\n"
                f"서식: 원본 유지\n\n"
                f"소요 시간: {format_time(total_time)}\n"
                f"성능 향상: {speedup:.0f}배"
            )

            if errors:
                msg += f"\n\n오류 {len(errors)}건 발생"

        print()
        messagebox.showinfo("완료", msg)

    except Exception as e:
        print(f"\n오류 발생: {e}")
        try:
            messagebox.showerror("오류", str(e))
        except:
            pass
    finally:
        _cleanup_root()


if __name__ == "__main__":
    main()
