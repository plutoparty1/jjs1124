from copy import copy
import os
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

IN_PATH = r"C:\Users\정산-PC\Desktop\jjs\25년11월_거래명세서_병합풀기_예시.xlsx"
OUT_PATH = r"C:\Users\정산-PC\Desktop\jjs\25년11월_거래명세서_병합없음_재작성.xlsx"

def safe_len(v) -> int:
    if v is None:
        return 0
    s = str(v)
    # 줄바꿈 포함이면 가장 긴 라인 기준
    return max((len(line) for line in s.splitlines()), default=0)

def calculate_column_width(text_len: int) -> float:
    """열 너비 계산 (한글 고려)"""
    # Excel 열 너비는 문자 수 기준
    # 한글/영문 혼합 시: 영문 1자 ≈ 1.1 units, 한글 1자 ≈ 2.0 units
    # 보수적으로 계산: text_len * 1.3 (한글 비율 고려)
    # 최소 너비 8, 여유 공간 2~3 추가
    if text_len == 0:
        return 8.0
    # 한글 비율을 고려한 너비 계산
    width = text_len * 1.3 + 3
    return max(8.0, width)

def copy_cell_style(src, dst):
    # openpyxl 스타일 객체는 immutable처럼 다뤄서 copy 필요
    dst.font = copy(src.font)
    dst.fill = copy(src.fill)
    dst.border = copy(src.border)
    dst.alignment = copy(src.alignment)
    dst.number_format = src.number_format
    dst.protection = copy(src.protection)
    dst.comment = copy(src.comment) if src.comment else None

def unmerge_and_fill(ws):
    # 병합 목록은 unmerge하면서 바뀌므로 미리 리스트로 고정
    merged_ranges = list(ws.merged_cells.ranges)

    for mrange in merged_ranges:
        min_row, min_col, max_row, max_col = mrange.min_row, mrange.min_col, mrange.max_row, mrange.max_col

        top_left = ws.cell(min_row, min_col)
        value = top_left.value

        # 병합 해제
        ws.unmerge_cells(str(mrange))

        # 값은 왼쪽 상단 셀에만 유지하고, 나머지 셀은 비움
        # 스타일은 전체 영역에 복사 (시각적 일관성 유지)
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                cell = ws.cell(r, c)
                # 왼쪽 상단 셀에만 값 유지
                if r == min_row and c == min_col:
                    cell.value = value
                else:
                    cell.value = None  # 나머지 셀은 비움
                # 스타일은 전체 영역에 복사
                copy_cell_style(top_left, cell)

def find_table_region(ws):
    """테이블 영역 찾기 (연속된 데이터가 있는 영역)"""
    if ws.max_row == 0 or ws.max_column == 0:
        return None, None, None, None
    
    # 각 행의 데이터 개수 계산
    row_data_count = {}
    for row_idx in range(1, ws.max_row + 1):
        data_count = 0
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row_idx, col_idx)
            if cell.value is not None and str(cell.value).strip() != "":
                data_count += 1
        row_data_count[row_idx] = data_count
    
    # 테이블 시작 행 찾기 (연속적으로 데이터가 많은 행)
    # 최소 2개 이상의 열에 데이터가 있는 행을 테이블로 간주
    table_start_row = None
    min_data_cols = 2  # 최소 2개 열에 데이터가 있어야 테이블로 간주
    
    for row_idx in range(1, ws.max_row + 1):
        if row_data_count[row_idx] >= min_data_cols:
            table_start_row = row_idx
            break
    
    if table_start_row is None:
        return None, None, None, None
    
    # 테이블 끝 행 찾기 (연속된 데이터 행의 마지막)
    table_end_row = table_start_row
    for row_idx in range(table_start_row + 1, ws.max_row + 1):
        if row_data_count[row_idx] >= min_data_cols:
            table_end_row = row_idx
        elif row_data_count[row_idx] == 0:
            # 빈 행이 나오면 테이블 종료로 간주
            break
    
    # 테이블 열 범위 찾기 (데이터가 있는 열)
    table_start_col = None
    table_end_col = None
    
    for row_idx in range(table_start_row, table_end_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row_idx, col_idx)
            if cell.value is not None and str(cell.value).strip() != "":
                if table_start_col is None:
                    table_start_col = col_idx
                table_start_col = min(table_start_col, col_idx)
                table_end_col = max(table_end_col, col_idx) if table_end_col else col_idx
    
    if table_start_col is None:
        return None, None, None, None
    
    return table_start_row, table_end_row, table_start_col, table_end_col

def move_data_to_table(ws, table_start_row, table_start_col, table_end_col):
    """테이블 위쪽의 데이터를 테이블 영역으로 이동"""
    if table_start_row <= 1:
        return  # 테이블이 첫 번째 행이면 이동할 데이터 없음
    
    moved_count = 0
    # 테이블 위쪽 행들을 역순으로 확인 (아래에서 위로)
    for row_idx in range(table_start_row - 1, 0, -1):
        row_has_data = False
        row_data = {}
        
        # 현재 행의 데이터 수집
        for col_idx in range(table_start_col, table_end_col + 1):
            cell = ws.cell(row_idx, col_idx)
            if cell.value is not None and str(cell.value).strip() != "":
                row_has_data = True
                row_data[col_idx] = {
                    'value': cell.value,
                    'style': cell
                }
        
        if not row_has_data:
            continue
        
        # 테이블 영역에서 빈 행 찾기
        moved = False
        for target_row in range(table_start_row, table_start_row + 5):  # 최대 5행까지만 확인
            if target_row > ws.max_row:
                break
            
            # 해당 행이 비어있는지 확인
            is_empty = True
            for col_idx in range(table_start_col, table_end_col + 1):
                cell = ws.cell(target_row, col_idx)
                if cell.value is not None and str(cell.value).strip() != "":
                    is_empty = False
                    break
            
            if is_empty:
                # 데이터 이동
                for col_idx, data in row_data.items():
                    src_cell = ws.cell(row_idx, col_idx)
                    dst_cell = ws.cell(target_row, col_idx)
                    dst_cell.value = data['value']
                    copy_cell_style(src_cell, dst_cell)
                    # 원본 셀 비우기
                    src_cell.value = None
                
                moved = True
                moved_count += 1
                break
        
        if moved:
            # 이동한 행 삭제 (빈 행이 되었으므로)
            try:
                ws.delete_rows(row_idx)
            except:
                pass
    
    if moved_count > 0:
        print(f"  테이블 위쪽 데이터 {moved_count}개 행을 테이블 영역으로 이동")

def clean_empty_cells_in_table(ws, table_start_row, table_end_row, table_start_col, table_end_col):
    """테이블 영역 내에서 빈 셀 정리 (위로 당기기)"""
    cleaned_count = 0
    
    # 각 열별로 처리
    for col_idx in range(table_start_col, table_end_col + 1):
        # 아래에서 위로 스캔하여 빈 셀을 찾고 위로 당기기
        for row_idx in range(table_end_row, table_start_row - 1, -1):
            cell = ws.cell(row_idx, col_idx)
            if cell.value is None or str(cell.value).strip() == "":
                # 위쪽에서 데이터가 있는 셀 찾기
                for search_row in range(row_idx - 1, table_start_row - 1, -1):
                    search_cell = ws.cell(search_row, col_idx)
                    if search_cell.value is not None and str(search_cell.value).strip() != "":
                        # 데이터 이동
                        cell.value = search_cell.value
                        copy_cell_style(search_cell, cell)
                        search_cell.value = None
                        cleaned_count += 1
                        break
    
    if cleaned_count > 0:
        print(f"  테이블 영역 내 빈 셀 {cleaned_count}개 정리 완료")

def remove_empty_columns(ws):
    """데이터가 없는 열을 삭제 (값이 실제로 있는 열만 유지)"""
    # 실제 사용된 열 찾기 (값이 있는 셀만)
    used_cols = set()
    max_used_col = 0
    
    for row in ws.iter_rows():
        for cell in row:
            # 값이 실제로 있는 셀만 확인 (None이 아니고 빈 문자열도 아님)
            if cell.value is not None:
                val_str = str(cell.value).strip()
                if val_str != "":
                    used_cols.add(cell.column)
                    max_used_col = max(max_used_col, cell.column)
    
    if not used_cols:
        # 모든 열이 비어있으면 첫 번째 열만 남김
        if ws.max_column > 1:
            ws.delete_cols(2, ws.max_column - 1)
            print(f"  모든 열이 비어있어 첫 번째 열만 유지")
        return
    
    # 사용되지 않은 열 찾기 (실제 사용된 최대 열까지 확인)
    empty_cols = []
    for col in range(1, max_used_col + 1):
        if col not in used_cols:
            empty_cols.append(col)
    
    # 뒤에서부터 삭제 (인덱스 꼬임 방지)
    deleted_count = 0
    for col in reversed(empty_cols):
        try:
            ws.delete_cols(col)
            deleted_count += 1
        except Exception as e:
            print(f"  열 {get_column_letter(col)} 삭제 실패: {e}")
    
    if deleted_count > 0:
        print(f"  빈 열 {deleted_count}개 삭제 완료")

def improve_readability(ws, max_width=45):
    # 전체 셀에 기본 정렬/줄바꿈 적용(너무 과하지 않게)
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None and cell.has_style is False:
                continue
            # 기존 alignment가 있으면 wrap만 켜고 나머지는 최대한 유지
            al = copy(cell.alignment)
            al.wrap_text = True
            # 세로 가운데, 가로는 기존 유지(없으면 left)
            if al.vertical is None:
                al.vertical = "center"
            if al.horizontal is None:
                al.horizontal = "left"
            cell.alignment = al

    # 열너비 “적당히” 자동 조정 (너무 넓어지는 거 방지)
    col_max = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            col = cell.column
            # 셀 내용 길이 계산
            cell_len = safe_len(cell.value)
            col_max[col] = max(col_max.get(col, 0), cell_len)

    # 각 열의 너비를 데이터 크기에 맞게 설정
    for col, text_len in col_max.items():
        col_letter = get_column_letter(col)
        # 한글 고려한 너비 계산
        width = calculate_column_width(text_len)
        # 최대 너비 제한
        width = min(width, max_width)
        ws.column_dimensions[col_letter].width = width
    
    # 데이터가 있지만 너비가 설정되지 않은 열은 기본 너비로 설정
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                col_letter = get_column_letter(cell.column)
                if col_letter not in ws.column_dimensions or ws.column_dimensions[col_letter].width is None:
                    text_len = safe_len(cell.value)
                    ws.column_dimensions[col_letter].width = calculate_column_width(text_len)

    # 첫 행 고정(원하면 바꿔도 됨)
    ws.freeze_panes = "A2"

def main():
    try:
        # 입력 파일 존재 확인
        if not os.path.exists(IN_PATH):
            print(f"오류: 입력 파일을 찾을 수 없습니다: {IN_PATH}")
            return
        
        # 출력 디렉토리 확인 및 생성
        out_dir = os.path.dirname(OUT_PATH)
        if out_dir and not os.path.exists(out_dir):
            os.makedirs(out_dir, exist_ok=True)
        
        # 워크북 로드
        print(f"파일 로딩 중: {IN_PATH}")
        wb = load_workbook(IN_PATH)
        
        if not wb.worksheets:
            print("경고: 워크시트가 없습니다.")
            return

        # 각 워크시트 처리
        for ws in wb.worksheets:
            print(f"처리 중: {ws.title}")
            # 1) 병합 완전 제거 + 값/서식 보존
            if ws.merged_cells.ranges:
                unmerge_and_fill(ws)

            # 2) 테이블 영역 찾기
            table_start_row, table_end_row, table_start_col, table_end_col = find_table_region(ws)
            
            if table_start_row:
                print(f"  테이블 영역: 행 {table_start_row}-{table_end_row}, 열 {get_column_letter(table_start_col)}-{get_column_letter(table_end_col)}")
                
                # 3) 테이블 위쪽 데이터를 테이블 영역으로 이동
                move_data_to_table(ws, table_start_row, table_start_col, table_end_col)
                
                # 테이블 영역 재계산 (이동 후 변경될 수 있음)
                table_start_row, table_end_row, table_start_col, table_end_col = find_table_region(ws)
                
                if table_start_row:
                    # 4) 테이블 영역 내 빈 셀 정리
                    clean_empty_cells_in_table(ws, table_start_row, table_end_row, table_start_col, table_end_col)
            else:
                print(f"  테이블 영역을 찾을 수 없습니다.")

            # 5) 빈 열 삭제
            remove_empty_columns(ws)

            # 6) 가독성 개선(줄바꿈/열너비/틀고정)
            improve_readability(ws)

        # 저장
        wb.save(OUT_PATH)
        print(f"완료: {OUT_PATH}")
        
    except PermissionError:
        print(f"오류: 파일 접근 권한이 없습니다. 파일이 다른 프로그램에서 열려있을 수 있습니다.")
    except Exception as e:
        print(f"오류 발생: {type(e).__name__}: {e}")
        raise

if __name__ == "__main__":
    main()
