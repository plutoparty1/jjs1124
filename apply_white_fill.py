"""
Excel 파일 복사 및 전체 셀에 흰색 배경 적용 스크립트
"""

from pathlib import Path
from typing import Optional, Tuple
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Font, Alignment, Side
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet
import logging
import win32com.client
import tkinter as tk
from tkinter import filedialog
import tempfile
import shutil
import os

# ============================================================================
# 상수 정의
# ============================================================================
MIN_CHECK_ROWS = 200
MIN_CHECK_COLS = 200
COLUMN_WIDTH_PIXELS = 80
COLUMN_WIDTH_CHARS = 11.43  # 80픽셀 ≈ 11.43 문자
TABLE_ROW_HEIGHT_PIXELS = 25
TABLE_ROW_HEIGHT_POINTS = 18.75  # 25픽셀 ≈ 18.75 포인트
IMAGE_MAX_WIDTH = 200
PIXELS_PER_CHAR = 7  # Excel 기준: 1 문자 ≈ 7 픽셀
EXCEL_DEFAULT_COLUMN_WIDTH = 8.43

# 폰트 설정
DEFAULT_FONT_NAME = '굴림'
DEFAULT_FONT_SIZE = 10
SPECIAL_FONT_NAME = 'HY헤드라인M'
SPECIAL_FONT_SIZE = 16

# 특수 텍스트 및 폰트 매핑
SPECIAL_TEXT_FONTS = {
    "주식회사 플랜티엠": Font(name=SPECIAL_FONT_NAME, size=SPECIAL_FONT_SIZE, bold=True),
    "대표이사 김 진 해": Font(name=SPECIAL_FONT_NAME, size=SPECIAL_FONT_SIZE, bold=True),
}

# 기본 범위
DEFAULT_RANGE = "B34:AL38"

# ============================================================================
# 로깅 설정
# ============================================================================
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S"
)
logger = logging.getLogger(__name__)


# ============================================================================
# 유틸리티 함수
# ============================================================================
def convert_xls_to_xlsx(xls_path: str) -> str:
    """
    .xls 파일을 .xlsx로 변환 (임시 파일 사용)
    
    Args:
        xls_path: .xls 파일 경로
        
    Returns:
        변환된 .xlsx 파일 경로
    """
    xls_path_obj = Path(xls_path)
    if not xls_path_obj.exists():
        raise FileNotFoundError(f"파일을 찾을 수 없습니다: {xls_path}")
    
    # 임시 디렉토리와 파일 경로 생성
    tmp_dir = tempfile.gettempdir()
    base_name = xls_path_obj.stem
    tmp_xls_path = os.path.join(tmp_dir, f"{base_name}_temp.xls")
    tmp_xlsx_path = os.path.join(tmp_dir, f"{base_name}_temp.xlsx")
    
    # 고유한 파일명 생성 (이미 존재하는 경우)
    counter = 1
    while os.path.exists(tmp_xls_path) or os.path.exists(tmp_xlsx_path):
        tmp_xls_path = os.path.join(tmp_dir, f"{base_name}_temp_{counter}.xls")
        tmp_xlsx_path = os.path.join(tmp_dir, f"{base_name}_temp_{counter}.xlsx")
        counter += 1
    
    try:
        # 원본 파일을 임시 디렉토리로 복사
        shutil.copy2(xls_path, tmp_xls_path)
        
        # Excel COM을 사용하여 변환
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = None
        
        try:
            wb = excel.Workbooks.Open(tmp_xls_path, ReadOnly=True)
            wb.SaveAs(tmp_xlsx_path, FileFormat=51)  # 51 = xlsx
            wb.Close(False)
            wb = None
        except Exception as e:
            raise RuntimeError(f".xls → .xlsx 변환 중 오류: {e}")
        finally:
            # 리소스 정리
            if wb is not None:
                try:
                    wb.Close(False)
                except Exception:
                    pass
            try:
                excel.Quit()
            except Exception:
                pass
            try:
                import pythoncom
                pythoncom.CoUninitialize()
            except Exception:
                pass
        
        logger.info(f"  - .xls 파일을 .xlsx로 변환 완료: {tmp_xlsx_path}")
        return tmp_xlsx_path
        
    finally:
        # 임시 .xls 파일 삭제
        try:
            if os.path.exists(tmp_xls_path):
                os.remove(tmp_xls_path)
        except (OSError, PermissionError):
            pass


def ensure_xlsx_file(input_file: str) -> Tuple[str, bool]:
    """
    입력 파일이 .xls인 경우 .xlsx로 변환하여 반환
    
    Args:
        input_file: 입력 파일 경로
        
    Returns:
        (파일 경로, 임시 파일 여부) 튜플
    """
    input_path = Path(input_file)
    is_temp = False
    
    if input_path.suffix.lower() == '.xls':
        logger.info(f"  - .xls 파일 감지: {input_file}")
        logger.info(f"  - .xlsx로 변환 중...")
        converted_path = convert_xls_to_xlsx(input_file)
        is_temp = True
        return converted_path, is_temp
    else:
        return str(input_path), is_temp
def get_sheet_dimensions(ws: Worksheet, min_rows: int = MIN_CHECK_ROWS, min_cols: int = MIN_CHECK_COLS) -> Tuple[int, int]:
    """
    시트의 최대 행/열을 가져오되 최소값 보장
    
    Args:
        ws: 워크시트 객체
        min_rows: 최소 행 수
        min_cols: 최소 열 수
        
    Returns:
        (check_rows, check_cols) 튜플
    """
    max_row = ws.max_row or 1
    max_column = ws.max_column or 1
    return max(max_row, min_rows), max(max_column, min_cols)


def parse_range(range_str: str) -> Tuple[int, int, int, int]:
    """
    범위 문자열을 파싱하여 행/열 번호 반환
    
    Args:
        range_str: 범위 문자열 (예: "B34:AL38")
        
    Returns:
        (start_row, end_row, start_col, end_col) 튜플
        
    Raises:
        ValueError: 유효하지 않은 범위 형식인 경우
    """
    if ':' not in range_str:
        raise ValueError(f"유효하지 않은 범위 형식입니다: {range_str}. 범위는 '시작셀:끝셀' 형식이어야 합니다 (예: B34:AL38)")
    
    start_cell, end_cell = range_str.split(':')
    
    start_col_str = ''.join(filter(str.isalpha, start_cell))
    start_row = int(''.join(filter(str.isdigit, start_cell)))
    start_col = column_index_from_string(start_col_str)
    
    end_col_str = ''.join(filter(str.isalpha, end_cell))
    end_row = int(''.join(filter(str.isdigit, end_cell)))
    end_col = column_index_from_string(end_col_str)
    
    return start_row, end_row, start_col, end_col


def safe_cell_access(ws: Worksheet, row: int, col: int, default=None):
    """
    안전하게 셀에 접근 (MergedCell 예외 처리)
    
    Args:
        ws: 워크시트 객체
        row: 행 번호
        col: 열 번호
        default: 예외 발생 시 반환할 기본값
        
    Returns:
        셀 객체 또는 default
    """
    try:
        return ws.cell(row=row, column=col)
    except (AttributeError, ValueError) as e:
        logger.debug(f"셀 접근 실패 ({row}, {col}): {e}")
        return default


def clear_cell_value(cell) -> bool:
    """
    셀 값을 안전하게 지움
    
    Args:
        cell: 셀 객체
        
    Returns:
        성공 여부
    """
    try:
        cell.value = None
        return True
    except (AttributeError, ValueError) as e:
        logger.debug(f"셀 값 지우기 실패: {e}")
        return False


# ============================================================================
# 주요 처리 함수
# ============================================================================
def apply_white_fill_to_all_cells(input_file: str, output_file: Optional[str] = None) -> Path:
    """
    Excel 파일을 복사하고 모든 셀에 흰색 배경을 적용합니다.
    숨겨진 행/열을 표시하고 데이터를 삭제합니다.
    
    Args:
        input_file: 원본 Excel 파일 경로
        output_file: 출력 파일 경로 (None이면 자동 생성)
        
    Returns:
        출력 파일 경로
    """
    input_path = Path(input_file)
    
    if not input_path.exists():
        raise FileNotFoundError(f"파일을 찾을 수 없습니다: {input_file}")
    
    # 출력 파일명 생성
    if output_file is None:
        stem = input_path.stem
        suffix = input_path.suffix
        output_path = input_path.parent / f"{stem}_흰색배경{suffix}"
    else:
        output_path = Path(output_file)
    
    logger.info(f"원본 파일 로드 중: {input_path}")
    
    # .xls 파일인 경우 .xlsx로 변환
    actual_file, is_temp = ensure_xlsx_file(str(input_path))
    temp_file_path = None
    if is_temp:
        temp_file_path = actual_file
    
    try:
        wb = load_workbook(actual_file)
    finally:
        # 임시 파일이면 삭제
        if is_temp and temp_file_path and os.path.exists(temp_file_path):
            try:
                os.remove(temp_file_path)
                logger.info(f"  - 임시 변환 파일 삭제 완료")
            except (OSError, PermissionError):
                logger.warning(f"  - 임시 변환 파일 삭제 실패: {temp_file_path}")
    
    # 모든 시트 처리
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        logger.info(f"시트 '{sheet_name}' 처리 중...")
        
        # 숨겨진 행/열 표시 및 데이터 지우기
        hidden_rows, hidden_cols = unhide_and_clear_hidden_cells(ws)
        
        if hidden_rows or hidden_cols:
            logger.info(f"  - 숨겨진 행 {len(hidden_rows)}개, 열 {len(hidden_cols)}개 표시 및 데이터 삭제 완료")
    
    # 파일 저장
    logger.info(f"파일 저장 중: {output_path}")
    wb.save(output_path)
    logger.info(f"완료! 저장된 파일: {output_path}")
    
    return output_path


def unhide_and_clear_hidden_cells(ws: Worksheet) -> Tuple[list, list]:
    """
    숨겨진 행/열을 표시하고 데이터를 지움
    
    Args:
        ws: 워크시트 객체
        
    Returns:
        (hidden_rows, hidden_cols) 튜플
    """
    hidden_rows = []
    hidden_cols = []
    
    # 숨겨진 행 찾기 및 표시
    for row_num in ws.row_dimensions:
        if ws.row_dimensions[row_num].hidden:
            ws.row_dimensions[row_num].hidden = False
            hidden_rows.append(row_num)
            # 숨겨진 행의 모든 셀 데이터 지우기
            max_col = ws.max_column or 1
            for col in range(1, max_col + 1):
                cell = safe_cell_access(ws, row_num, col)
                if cell:
                    clear_cell_value(cell)
    
    # 숨겨진 열 찾기 및 표시
    max_col = ws.max_column or 1
    check_cols = max(max_col + 100, MIN_CHECK_COLS)
    
    logger.info(f"  - 열 확인 범위: 1 ~ {check_cols} (최대 열: {max_col})")
    
    for col_num in range(1, check_cols + 1):
        col_letter = get_column_letter(col_num)
        col_dim = ws.column_dimensions[col_letter]
        
        hidden_value = getattr(col_dim, 'hidden', None)
        
        if col_num == 3:
            logger.info(f"    - C열 확인: hidden={hidden_value}, type={type(hidden_value)}")
        
        if hidden_value is True:
            col_dim.hidden = False
            hidden_cols.append(col_letter)
            logger.info(f"    - 숨겨진 열 발견 및 표시: {col_letter} (열 {col_num})")
            
            # 숨겨진 열의 모든 셀 데이터 지우기
            max_row = ws.max_row or 1
            for row in range(1, max_row + 1):
                cell = safe_cell_access(ws, row, col_num)
                if cell:
                    clear_cell_value(cell)
        elif hidden_value is not None and hidden_value is not False:
            logger.info(f"    - 열 {col_letter}의 hidden 속성이 예상치 못한 값: {hidden_value}")
            col_dim.hidden = False
    
    return hidden_rows, hidden_cols


def unmerge_cells_in_table_range(ws: Worksheet, start_row: int, end_row: int, start_col: int, end_col: int) -> list:
    """
    테이블 범위 내의 병합 셀을 해제
    
    Args:
        ws: 워크시트 객체
        start_row: 시작 행
        end_row: 끝 행
        start_col: 시작 열
        end_col: 끝 열
        
    Returns:
        해제된 병합 범위 리스트
    """
    merged_ranges = list(ws.merged_cells.ranges)
    ranges_to_unmerge = []
    
    for merged_range in merged_ranges:
        if (merged_range.min_row <= end_row and merged_range.max_row >= start_row and
            merged_range.min_col <= end_col and merged_range.max_col >= start_col):
            ranges_to_unmerge.append(str(merged_range))
    
    for merge_range in ranges_to_unmerge:
        ws.unmerge_cells(merge_range)
        logger.info(f"  - 병합 해제: {merge_range}")
    
    return ranges_to_unmerge


def rearrange_data_in_table_range(ws: Worksheet, start_row: int, end_row: int, start_col: int, end_col: int) -> int:
    """
    테이블 범위 내의 데이터를 재배치 (빈 셀 제거, 왼쪽으로 붙이기)
    
    Args:
        ws: 워크시트 객체
        start_row: 시작 행
        end_row: 끝 행
        start_col: 시작 열
        end_col: 끝 열
        
    Returns:
        정리된 테이블의 마지막 열 번호
    """
    table_last_col = start_col - 1
    
    for row in range(start_row, end_row + 1):
        # 해당 행의 데이터 수집 (빈 셀 제외)
        row_data = []
        for col in range(start_col, end_col + 1):
            cell = safe_cell_access(ws, row, col)
            if cell and cell.value is not None and str(cell.value).strip() != '':
                row_data.append(cell.value)
        
        # 기존 범위의 셀 모두 지우기
        for col in range(start_col, end_col + 1):
            cell = safe_cell_access(ws, row, col)
            if cell:
                clear_cell_value(cell)
        
        # 데이터를 B열부터 왼쪽으로 붙이기
        for idx, value in enumerate(row_data):
            col = start_col + idx
            cell = safe_cell_access(ws, row, col)
            if cell:
                cell.value = value
                table_last_col = max(table_last_col, col)
        
        logger.info(f"  - 행 {row}: {len(row_data)}개 데이터를 B열부터 재배치 완료")
    
    logger.info(f"  - 정리된 테이블의 마지막 열: {get_column_letter(table_last_col)} (열 {table_last_col})")
    return table_last_col


def process_merged_cells_outside_table(
    ws: Worksheet, 
    start_row: int, 
    end_row: int, 
    start_col: int, 
    end_col: int
) -> set:
    """
    테이블 범위 밖의 병합 셀을 해제하고 데이터를 B열로 이동
    
    Args:
        ws: 워크시트 객체
        start_row: 테이블 시작 행
        end_row: 테이블 끝 행
        start_col: 테이블 시작 열
        end_col: 테이블 끝 열
        
    Returns:
        B열로 데이터가 이동된 행 번호 집합
    """
    rows_with_merged_data_moved = set()
    merged_cells_info = []
    
    # 병합 셀 분류
    a_col_merged = []
    other_merged = []
    next_row_merged = []
    
    for merged_range in list(ws.merged_cells.ranges):
        if not (merged_range.min_row <= end_row and merged_range.max_row >= start_row and
                merged_range.min_col <= end_col and merged_range.max_col >= start_col):
            
            if merged_range.min_col == 1:
                a_col_merged.append(merged_range)
                merged_cells_info.append({
                    'type': 'A열 병합 셀 (해제됨, B열로 이동)',
                    'range': str(merged_range),
                    'row': f"{merged_range.min_row}-{merged_range.max_row}"
                })
            elif merged_range.min_col >= start_col:
                next_row_merged.append(merged_range)
                merged_cells_info.append({
                    'type': '병합 셀 (해제됨, B열로 이동)',
                    'range': str(merged_range),
                    'row': f"{merged_range.min_row}-{merged_range.max_row}"
                })
            elif merged_range.min_col < start_col and merged_range.max_col >= start_col:
                other_merged.append(merged_range)
                merged_cells_info.append({
                    'type': 'A-C열 병합 셀 (해제됨, B열로 이동)',
                    'range': str(merged_range),
                    'row': f"{merged_range.min_row}-{merged_range.max_row}"
                })
    
    # 각 타입별로 병합 셀 처리
    for merged_range in other_merged:
        rows_with_merged_data_moved.update(
            unmerge_and_move_to_b_col(ws, merged_range, start_col, "A-C열")
        )
    
    for merged_range in a_col_merged:
        rows_with_merged_data_moved.update(
            unmerge_and_move_to_b_col(ws, merged_range, start_col, "A열")
        )
    
    for merged_range in next_row_merged:
        rows_with_merged_data_moved.update(
            unmerge_and_move_to_b_col(ws, merged_range, start_col, "다른 열")
        )
    
    if merged_cells_info:
        logger.info(f"  - 처리된 병합 셀: {len(merged_cells_info)}개")
        for info in merged_cells_info:
            logger.info(f"    * {info['type']}: {info['range']} (행 {info['row']})")
    
    return rows_with_merged_data_moved


def unmerge_and_move_to_b_col(
    ws: Worksheet, 
    merged_range, 
    start_col: int, 
    range_type: str
) -> set:
    """
    병합 셀을 해제하고 데이터를 B열로 이동
    
    Args:
        ws: 워크시트 객체
        merged_range: 병합 범위 객체
        start_col: 시작 열 (B열)
        range_type: 범위 타입 설명
        
    Returns:
        데이터가 이동된 행 번호 집합
    """
    rows_moved = set()
    
    try:
        origin_row = merged_range.min_row
        origin_col = merged_range.min_col
        origin_cell = safe_cell_access(ws, origin_row, origin_col)
        
        if not origin_cell:
            return rows_moved
        
        # 병합 해제 전에 데이터 저장
        merged_data = None
        try:
            merged_data = origin_cell.value
        except (AttributeError, ValueError):
            pass
        
        # B열에 원래 있던 데이터 저장
        original_b_col_data = {}
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            b_cell = safe_cell_access(ws, row, start_col)
            if b_cell and b_cell.value is not None and str(b_cell.value).strip() != '':
                original_b_col_data[row] = b_cell.value
        
        # 병합 해제
        ws.unmerge_cells(str(merged_range))
        logger.info(f"  - {range_type} 병합 해제: {merged_range}")
        
        # 데이터가 있으면 B열로 이동
        if merged_data is not None and str(merged_data).strip() != '':
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                # 병합 범위 내의 모든 셀 데이터 지우기
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    cell = safe_cell_access(ws, row, col)
                    if cell:
                        clear_cell_value(cell)
                
                # B열에 데이터 배치
                b_cell = safe_cell_access(ws, row, start_col)
                if b_cell:
                    if row in original_b_col_data:
                        b_cell.value = original_b_col_data[row]
                        logger.info(f"    - 행 {row}: 원래 B열 데이터 유지 ({get_column_letter(origin_col)}열 병합 데이터는 배치하지 않음)")
                    else:
                        b_cell.value = merged_data
                        logger.info(f"    - 행 {row}: {get_column_letter(origin_col)}열 데이터를 B열로 이동")
                        rows_moved.add(row)
    except Exception as e:
        logger.warning(f"  - {range_type} 병합 해제 실패: {merged_range} - {e}")
    
    return rows_moved


def process_other_rows(
    ws: Worksheet,
    all_rows_to_process: list,
    start_col: int,
    end_col: int,
    rows_with_merged_data_moved: set
) -> list:
    """
    테이블 범위 밖의 다른 행들을 처리
    
    Args:
        ws: 워크시트 객체
        all_rows_to_process: 처리할 행 번호 리스트
        start_col: 시작 열
        end_col: 끝 열
        rows_with_merged_data_moved: B열로 데이터가 이동된 행 집합
        
    Returns:
        처리된 행 번호 리스트
    """
    processed_rows = []
    check_cols = ws.max_column + 100 if ws.max_column else end_col + 200
    
    for row in all_rows_to_process:
        row_data = []
        
        # 병합 셀 해제 후 B열로 이동한 행인 경우, B열 데이터를 먼저 수집
        if row in rows_with_merged_data_moved:
            b_cell = safe_cell_access(ws, row, start_col)
            if b_cell and b_cell.value is not None and str(b_cell.value).strip() != '':
                row_data.append(b_cell.value)
        
        # A열 확인 및 데이터 수집
        if row not in rows_with_merged_data_moved:
            a_cell = safe_cell_access(ws, row, 1)
            if a_cell and a_cell.value is not None and str(a_cell.value).strip() != '':
                row_data.append(a_cell.value)
        
        # B열 이후의 데이터 수집
        start_col_for_collection = start_col + 1 if row in rows_with_merged_data_moved else start_col
        for col in range(start_col_for_collection, check_cols):
            cell = safe_cell_access(ws, row, col)
            if cell and cell.value is not None and str(cell.value).strip() != '':
                row_data.append(cell.value)
        
        if not row_data:
            continue
        
        # 기존 행의 데이터 지우기
        if row not in rows_with_merged_data_moved:
            a_cell = safe_cell_access(ws, row, 1)
            if a_cell:
                clear_cell_value(a_cell)
        
        start_col_for_clear = start_col + 1 if row in rows_with_merged_data_moved else start_col
        for col in range(start_col_for_clear, check_cols):
            cell = safe_cell_access(ws, row, col)
            if cell:
                clear_cell_value(cell)
        
        # 데이터를 B열부터 왼쪽으로 붙이기
        for idx, value in enumerate(row_data):
            col = start_col + idx
            cell = safe_cell_access(ws, row, col)
            if cell:
                cell.value = value
        
        processed_rows.append(row)
        logger.info(f"  - 행 {row}: {len(row_data)}개 데이터를 B열부터 재배치 완료")
    
    return processed_rows


def apply_white_fill_to_all_sheets(wb) -> None:
    """
    모든 시트의 모든 셀에 흰색 채우기 적용
    
    Args:
        wb: 워크북 객체
    """
    logger.info(f"\n[마지막 단계] 모든 셀에 흰색 채우기 적용 중...")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        logger.info(f"  - 시트 '{sheet_name}' 처리 중...")
        
        check_rows, check_cols = get_sheet_dimensions(ws)
        
        cell_count = 0
        for row in range(1, check_rows + 1):
            for col in range(1, check_cols + 1):
                cell = safe_cell_access(ws, row, col)
                if cell:
                    cell.fill = white_fill
                    cell_count += 1
        
        logger.info(f"    - {cell_count}개 셀에 흰색 배경 적용 완료")
    
    logger.info(f"[마지막 단계] 완료!")


def apply_white_fill_to_sheet(ws: Worksheet) -> None:
    """
    지정된 시트의 모든 셀에 흰색 채우기 적용
    
    Args:
        ws: 워크시트 객체
    """
    logger.info(f"\n[마지막 단계] 시트 '{ws.title}'의 모든 셀에 흰색 채우기 적용 중...")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    check_rows, check_cols = get_sheet_dimensions(ws)
    
    cell_count = 0
    for row in range(1, check_rows + 1):
        for col in range(1, check_cols + 1):
            cell = safe_cell_access(ws, row, col)
            if cell:
                try:
                    cell.fill = white_fill
                    cell_count += 1
                except Exception as e:
                    logger.debug(f"셀 채우기 실패 ({row}, {col}): {e}")
    
    logger.info(f"  - {cell_count}개 셀에 흰색 배경 적용 완료")
    logger.info(f"[마지막 단계] 완료!")


def replace_text_in_all_sheets(wb) -> None:
    """
    모든 시트에서 지정된 텍스트를 찾아서 바꿈
    
    Args:
        wb: 워크북 객체
    """
    # 찾기/바꾸기 매핑
    replacements = {
        "한민정": "정지수",
        "hmj0518": "jjs1124",
        "수    신 ": "수    신 : ",
        "7181.": "7181"
    }
    
    logger.info(f"\n[텍스트 찾기/바꾸기] 모든 시트에서 텍스트 교체 중...")
    
    total_replacements = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_replacements = 0
        
        check_rows, check_cols = get_sheet_dimensions(ws)
        
        for row in range(1, check_rows + 1):
            for col in range(1, check_cols + 1):
                cell = safe_cell_access(ws, row, col)
                if cell and cell.value:
                    # 셀 값이 문자열인 경우에만 처리
                    if isinstance(cell.value, str):
                        original_value = cell.value
                        new_value = original_value
                        
                        # 모든 찾기/바꾸기 적용
                        for old_text, new_text in replacements.items():
                            if old_text in new_value:
                                new_value = new_value.replace(old_text, new_text)
                                sheet_replacements += 1
                        
                        # 값이 변경된 경우에만 셀 업데이트
                        if new_value != original_value:
                            try:
                                cell.value = new_value
                            except Exception as e:
                                logger.debug(f"셀 값 업데이트 실패 ({row}, {col}): {e}")
        
        if sheet_replacements > 0:
            logger.info(f"  - 시트 '{sheet_name}': {sheet_replacements}개 텍스트 교체 완료")
            total_replacements += sheet_replacements
    
    if total_replacements > 0:
        logger.info(f"  - 총 {total_replacements}개 텍스트 교체 완료")
    else:
        logger.info(f"  - 교체할 텍스트 없음")
    
    logger.info(f"[텍스트 찾기/바꾸기] 완료!")


def apply_fonts_and_row_heights(
    ws: Worksheet,
    start_row: int,
    end_row: int
) -> None:
    """
    모든 셀에 폰트 적용 및 테이블 행 높이 설정
    
    Args:
        ws: 워크시트 객체
        start_row: 테이블 시작 행
        end_row: 테이블 끝 행
    """
    logger.info(f"\n[폰트 및 행 높이 설정] 작업 시트('{ws.title}')의 모든 셀에 굴림체 10pt 적용 및 테이블 행 높이 설정 중...")
    
    gulim_font = Font(name=DEFAULT_FONT_NAME, size=DEFAULT_FONT_SIZE)
    check_rows, check_cols = get_sheet_dimensions(ws)
    
    # 모든 셀에 굴림체 10pt 적용
    font_count = 0
    for row in range(1, check_rows + 1):
        for col in range(1, check_cols + 1):
            cell = safe_cell_access(ws, row, col)
            if cell:
                try:
                    if cell.font:
                        cell.font = Font(
                            name=DEFAULT_FONT_NAME,
                            size=DEFAULT_FONT_SIZE,
                            bold=cell.font.bold,
                            italic=cell.font.italic,
                            underline=cell.font.underline,
                            strike=cell.font.strike,
                            color=cell.font.color
                        )
                    else:
                        cell.font = gulim_font
                    font_count += 1
                except Exception as e:
                    logger.debug(f"폰트 적용 실패 ({row}, {col}): {e}")
    
    # 테이블 행 높이 설정
    row_height_count = 0
    for row in range(start_row, end_row + 1):
        try:
            ws.row_dimensions[row].height = TABLE_ROW_HEIGHT_POINTS
            row_height_count += 1
        except Exception as e:
            logger.debug(f"행 높이 설정 실패 ({row}): {e}")
    
    logger.info(f"  - {font_count}개 셀에 굴림체 10pt 적용 완료")
    logger.info(f"  - 테이블 행({start_row}~{end_row}) {row_height_count}개 행의 높이를 {TABLE_ROW_HEIGHT_POINTS}포인트(약 {TABLE_ROW_HEIGHT_PIXELS}픽셀)로 설정 완료")
    
    # 특정 텍스트에 특수 폰트 적용
    special_font_count = 0
    for row in range(1, check_rows + 1):
        for col in range(1, check_cols + 1):
            cell = safe_cell_access(ws, row, col)
            if cell and cell.value is not None:
                cell_value_str = str(cell.value).strip()
                for special_text, special_font in SPECIAL_TEXT_FONTS.items():
                    if special_text in cell_value_str:
                        try:
                            cell.font = special_font
                            special_font_count += 1
                            logger.info(f"  - 셀 {get_column_letter(col)}{row}: '{special_text}' 텍스트에 {SPECIAL_FONT_NAME} {SPECIAL_FONT_SIZE}pt Bold 적용")
                            break
                        except Exception as e:
                            logger.debug(f"특수 폰트 적용 실패 ({row}, {col}): {e}")
    
    logger.info(f"  - {special_font_count}개 셀에 {SPECIAL_FONT_NAME} {SPECIAL_FONT_SIZE}pt Bold 적용 완료")
    logger.info(f"[폰트 및 행 높이 설정] 완료!")


def remove_all_borders(ws: Worksheet) -> None:
    """
    모든 셀에서 테두리 제거
    
    Args:
        ws: 워크시트 객체
    """
    logger.info(f"\n[1단계: 모든 테두리 삭제] 작업 시트('{ws.title}')의 모든 셀에서 테두리 삭제 중...")
    
    check_rows, check_cols = get_sheet_dimensions(ws)
    no_border = Border()
    
    border_removed_count = 0
    for row in range(1, check_rows + 1):
        for col in range(1, check_cols + 1):
            cell = safe_cell_access(ws, row, col)
            if cell:
                try:
                    cell.border = no_border
                    border_removed_count += 1
                except Exception as e:
                    logger.debug(f"테두리 제거 실패 ({row}, {col}): {e}")
    
    logger.info(f"  - {border_removed_count}개 셀에서 테두리 삭제 완료")
    logger.info(f"[1단계: 모든 테두리 삭제] 완료!")


def apply_table_borders_and_styles(
    ws: Worksheet,
    start_row: int,
    end_row: int,
    start_col: int,
    end_col: int
) -> None:
    """
    테이블 범위에 테두리 및 스타일 적용
    
    Args:
        ws: 워크시트 객체
        start_row: 시작 행
        end_row: 끝 행
        start_col: 시작 열
        end_col: 끝 열
    """
    logger.info(f"\n[2단계: 테두리 및 스타일 적용] 작업 시트('{ws.title}')의 지정된 범위에서 데이터가 있는 셀에만 테두리 설정 및 맨 위/아래 행 스타일 적용 중...")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal='center', vertical='center')
    number_format = '#,###'
    
    border_count = 0
    style_count = 0
    alignment_count = 0
    format_count = 0
    
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = safe_cell_access(ws, row, col)
            if cell and cell.value is not None and str(cell.value).strip() != '':
                try:
                    cell.border = thin_border
                    cell.alignment = center_alignment
                    cell.number_format = number_format
                    border_count += 1
                    alignment_count += 1
                    format_count += 1
                    
                    if row == start_row or row == end_row:
                        cell.font = bold_font
                        cell.fill = gray_fill
                        style_count += 1
                except Exception as e:
                    logger.debug(f"스타일 적용 실패 ({row}, {col}): {e}")
    
    logger.info(f"  - {border_count}개 셀에 테두리 설정 완료")
    logger.info(f"  - {alignment_count}개 셀에 가운데 정렬 적용 완료")
    logger.info(f"  - {format_count}개 셀에 #,### 서식 적용 완료")
    logger.info(f"  - 맨 위 행({start_row})과 맨 아래 행({end_row})의 {style_count}개 셀에 Bold 및 회색 채우기 적용 완료")
    logger.info(f"[2단계: 테두리 및 스타일 적용] 완료!")


def set_table_column_widths(ws: Worksheet, start_col: int, end_col: int) -> None:
    """
    테이블 범위 내 모든 열의 너비를 고정 너비로 설정
    
    Args:
        ws: 워크시트 객체
        start_col: 시작 열
        end_col: 끝 열
    """
    logger.info(f"\n[3단계: 열 너비 고정] 작업 시트('{ws.title}')의 지정된 범위 내 모든 열의 너비를 {COLUMN_WIDTH_PIXELS}픽셀({COLUMN_WIDTH_CHARS} 문자)로 설정 중...")
    
    width_adjusted_count = 0
    for col in range(start_col, end_col + 1):
        try:
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = COLUMN_WIDTH_CHARS
            width_adjusted_count += 1
        except Exception as e:
            logger.debug(f"열 너비 설정 실패 ({col_letter}): {e}")
    
    logger.info(f"  - {width_adjusted_count}개 열의 너비를 {COLUMN_WIDTH_CHARS} 문자(약 {COLUMN_WIDTH_PIXELS}픽셀)로 설정 완료")
    logger.info(f"[3단계: 열 너비 고정] 완료!")


def place_image_at_center_top(
    ws: Worksheet,
    start_row: int,
    end_row: int,
    start_col: int,
    end_col: int
) -> None:
    """
    그림을 맨 윗줄에 테이블 기준 가운데에 배치
    
    Args:
        ws: 워크시트 객체
        start_row: 테이블 시작 행
        end_row: 테이블 끝 행
        start_col: 테이블 시작 열
        end_col: 테이블 끝 열
    """
    logger.info(f"\n[4단계: 그림 배치] 작업 시트('{ws.title}')의 그림을 맨 윗줄에 테이블 기준 가운데에 배치 중...")
    
    try:
        images = []
        if hasattr(ws, '_images') and ws._images:
            images = list(ws._images)
        
        if not images:
            logger.warning("  - 그림을 찾을 수 없습니다.")
            return
        
        img = images[0]
        ws._images.clear()
        
        # 테이블의 실제 데이터가 있는 마지막 열 찾기
        actual_last_col = start_col - 1
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = safe_cell_access(ws, row, col)
                if cell and cell.value is not None and str(cell.value).strip() != '':
                    actual_last_col = max(actual_last_col, col)
        
        # 픽셀 단위로 테이블 가운데 계산
        if actual_last_col >= start_col:
            total_pixels = 0
            col_pixels = {}
            
            for col in range(start_col, actual_last_col + 1):
                col_letter = get_column_letter(col)
                col_width = ws.column_dimensions[col_letter].width
                if col_width is None:
                    col_width = EXCEL_DEFAULT_COLUMN_WIDTH
                
                col_pixel_width = col_width * PIXELS_PER_CHAR
                total_pixels += col_pixel_width
                col_pixels[col] = total_pixels
            
            center_pixel = total_pixels / 2
            center_col = start_col
            
            for col in range(start_col, actual_last_col + 1):
                if col_pixels[col] >= center_pixel:
                    center_col = col
                    break
            
            logger.info(f"  - 테이블 데이터 범위: {get_column_letter(start_col)}열 ~ {get_column_letter(actual_last_col)}열")
            logger.info(f"  - 테이블 총 너비: {total_pixels:.1f}픽셀")
            logger.info(f"  - 계산된 가운데 위치: {center_pixel:.1f}픽셀")
            logger.info(f"  - 가운데 열: {get_column_letter(center_col)}열 (열 {center_col})")
        else:
            center_col = (start_col + end_col) // 2
            logger.info(f"  - 데이터가 없어 지정된 범위의 중간 열 사용: {get_column_letter(center_col)}열")
        
        center_col_letter = get_column_letter(center_col)
        anchor_cell = f"{center_col_letter}1"
        img.anchor = anchor_cell
        
        # 비율 유지하면서 크기 조절
        if hasattr(img, 'width') and img.width:
            original_width = img.width
            if original_width > IMAGE_MAX_WIDTH:
                if hasattr(img, 'height') and img.height:
                    aspect_ratio = img.height / original_width if original_width > 0 else 1
                    img.width = IMAGE_MAX_WIDTH
                    img.height = IMAGE_MAX_WIDTH * aspect_ratio
                else:
                    img.width = IMAGE_MAX_WIDTH
        
        ws.add_image(img)
        logger.info(f"  - 그림을 1행 {anchor_cell} 셀에 배치 완료 (테이블 기준 가운데)")
        logger.info(f"  - 비율 유지 상태로 크기 조정 완료")
    except Exception as e:
        logger.warning(f"  - 그림 배치 중 오류: {e}")
    
    logger.info(f"[4단계: 그림 배치] 완료!")


def compact_range_data(
    input_file: str, 
    range_str: str = DEFAULT_RANGE, 
    sheet_name: Optional[str] = None, 
    output_file: Optional[str] = None
) -> Path:
    """
    지정된 범위의 데이터를 처리합니다:
    1. 셀 병합 해제
    2. 빈 셀 제거
    3. 데이터를 왼쪽으로 붙이기 (B열부터)
    4. 서식 적용
    
    Args:
        input_file: 원본 Excel 파일 경로
        range_str: 처리할 범위 (예: "B34:AL38")
        sheet_name: 시트 이름 (None이면 첫 번째 시트)
        output_file: 출력 파일 경로 (None이면 자동 생성)
        
    Returns:
        출력 파일 경로
    """
    input_path = Path(input_file)
    
    if not input_path.exists():
        raise FileNotFoundError(f"파일을 찾을 수 없습니다: {input_file}")
    
    logger.info(f"원본 파일 로드 중: {input_path}")
    
    # .xls 파일인 경우 .xlsx로 변환
    actual_file, is_temp = ensure_xlsx_file(str(input_path))
    temp_file_path = None
    if is_temp:
        temp_file_path = actual_file
    
    try:
        wb = load_workbook(actual_file)
    finally:
        # 임시 파일이면 삭제
        if is_temp and temp_file_path and os.path.exists(temp_file_path):
            try:
                os.remove(temp_file_path)
                logger.info(f"  - 임시 변환 파일 삭제 완료")
            except (OSError, PermissionError):
                logger.warning(f"  - 임시 변환 파일 삭제 실패: {temp_file_path}")
    
    # 시트 선택
    if sheet_name is None:
        ws = wb.active
    else:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"시트를 찾을 수 없습니다: {sheet_name}")
        ws = wb[sheet_name]
    
    logger.info(f"작업 시트: {ws.title}")
    target_worksheet = ws
    
    # 범위 파싱
    start_row, end_row, start_col, end_col = parse_range(range_str)
    logger.info(f"처리 범위: {range_str} (행 {start_row}-{end_row}, 열 {start_col}-{end_col})")
    
    # 1. 테이블 범위 내 병합 셀 해제
    unmerge_cells_in_table_range(ws, start_row, end_row, start_col, end_col)
    
    # 2. 테이블 범위 내 데이터 재배치
    table_last_col = rearrange_data_in_table_range(ws, start_row, end_row, start_col, end_col)
    
    # 3. 테이블 범위 밖의 병합 셀 처리
    max_ws_row = ws.max_row or end_row
    rows_above = list(range(1, start_row))
    rows_below = list(range(end_row + 1, max_ws_row + 1))
    all_rows_to_process = rows_above + rows_below
    
    logger.info(f"  - 다른 행 처리 시작: 위쪽 {len(rows_above)}개 행 (1 ~ {start_row-1}), 아래쪽 {len(rows_below)}개 행 ({end_row+1} ~ {max_ws_row})")
    
    rows_with_merged_data_moved = process_merged_cells_outside_table(
        ws, start_row, end_row, start_col, end_col
    )
    
    # 4. 다른 행들 처리
    processed_rows = process_other_rows(
        ws, all_rows_to_process, start_col, end_col, rows_with_merged_data_moved
    )
    
    if processed_rows:
        logger.info(f"  - 다음 행 처리 완료: 총 {len(processed_rows)}개 행 (행 {min(processed_rows)} ~ {max(processed_rows)})")
    else:
        logger.info(f"  - 처리할 다음 행 없음")
    
    # 5. 모든 셀에 흰색 채우기 적용
    apply_white_fill_to_all_sheets(wb)
    
    # 6. 폰트 및 행 높이 설정
    apply_fonts_and_row_heights(ws, start_row, end_row)
    
    # 7. 모든 테두리 삭제
    remove_all_borders(ws)
    
    # 8. 테이블 테두리 및 스타일 적용
    apply_table_borders_and_styles(ws, start_row, end_row, start_col, end_col)
    
    # 9. 열 너비 설정
    set_table_column_widths(ws, start_col, end_col)
    
    # 10. 그림 배치
    place_image_at_center_top(ws, start_row, end_row, start_col, end_col)
    
    # 출력 파일명 생성
    if output_file is None:
        stem = input_path.stem
        suffix = input_path.suffix
        range_clean = range_str.replace(':', '_')
        output_path = input_path.parent / f"{stem}_{range_clean}_정리{suffix}"
    else:
        output_path = Path(output_file)
    
    logger.info(f"파일 저장 중: {output_path}")
    wb.save(output_path)
    logger.info(f"완료! 저장된 파일: {output_path}")
    
    return output_path


def process_excel_file(
    input_file: str, 
    range_str: str = DEFAULT_RANGE, 
    output_file: Optional[str] = None
) -> Path:
    """
    Excel 파일을 완전히 처리하는 통합 함수:
    - 숨겨진 행/열 표시 및 데이터 삭제
    - 지정된 범위 데이터 정리 (병합 해제, 빈 셀 제거, 왼쪽으로 붙이기)
    - 서식 적용 (흰색 배경, 테두리, 폰트 등)
    
    Args:
        input_file: 원본 Excel 파일 경로
        range_str: 처리할 범위 (예: "B34:AL38")
        output_file: 최종 출력 파일 경로 (None이면 자동 생성)
        
    Returns:
        최종 출력 파일 경로
    """
    input_path = Path(input_file)
    
    if not input_path.exists():
        raise FileNotFoundError(f"파일을 찾을 수 없습니다: {input_file}")
    
    logger.info("=" * 60)
    logger.info("Excel 파일 통합 처리 시작")
    logger.info("=" * 60)
    
    # 최종 출력 파일명 생성
    if output_file is None:
        stem = input_path.stem
        suffix = input_path.suffix
        # .xls 파일인 경우 출력은 .xlsx로 저장
        if suffix.lower() == '.xls':
            suffix = '.xlsx'
        output_path = input_path.parent / f"{stem}_수정완료{suffix}"
    else:
        output_path = Path(output_file)
        # 출력 파일 확장자가 .xls인 경우 .xlsx로 변경
        if output_path.suffix.lower() == '.xls':
            output_path = output_path.with_suffix('.xlsx')
    
    # 파일을 한 번만 로드하여 모든 처리 수행
    logger.info(f"원본 파일 로드 중: {input_path}")
    
    # .xls 파일인 경우 .xlsx로 변환
    actual_file, is_temp = ensure_xlsx_file(str(input_path))
    temp_file_path = None
    if is_temp:
        temp_file_path = actual_file
    
    try:
        wb = load_workbook(actual_file)
    finally:
        # 임시 파일이면 삭제
        if is_temp and temp_file_path and os.path.exists(temp_file_path):
            try:
                os.remove(temp_file_path)
                logger.info(f"  - 임시 변환 파일 삭제 완료")
            except (OSError, PermissionError):
                logger.warning(f"  - 임시 변환 파일 삭제 실패: {temp_file_path}")
    
    # 첫 번째 시트만 선택 (나머지 시트는 원본 그대로 유지)
    ws = wb.worksheets[0]  # 첫 번째 시트 (인덱스 0)
    logger.info(f"작업 시트: {ws.title} (첫 번째 시트만 처리, 나머지 시트는 원본 그대로 유지)")
    
    # 첫 번째 시트에서만 숨겨진 행/열 표시 및 데이터 삭제
    logger.info("\n[1단계] 숨겨진 행/열 표시 및 데이터 삭제 중...")
    hidden_rows, hidden_cols = unhide_and_clear_hidden_cells(ws)
    if hidden_rows or hidden_cols:
        logger.info(f"  - 숨겨진 행 {len(hidden_rows)}개, 열 {len(hidden_cols)}개 표시 및 데이터 삭제 완료")
    logger.info("[1단계] 완료!")
    
    # 지정된 범위 데이터 정리 및 서식 적용
    logger.info(f"\n[2단계] {range_str} 범위 데이터 정리 및 서식 적용 중...")
    target_worksheet = ws
    
    # 범위 파싱
    start_row, end_row, start_col, end_col = parse_range(range_str)
    logger.info(f"처리 범위: {range_str} (행 {start_row}-{end_row}, 열 {start_col}-{end_col})")
    
    # 1. 테이블 범위 내 병합 셀 해제
    unmerge_cells_in_table_range(ws, start_row, end_row, start_col, end_col)
    
    # 2. 테이블 범위 내 데이터 재배치
    table_last_col = rearrange_data_in_table_range(ws, start_row, end_row, start_col, end_col)
    
    # 3. 테이블 범위 밖의 병합 셀 처리
    max_ws_row = ws.max_row or end_row
    rows_above = list(range(1, start_row))
    rows_below = list(range(end_row + 1, max_ws_row + 1))
    all_rows_to_process = rows_above + rows_below
    
    logger.info(f"  - 다른 행 처리 시작: 위쪽 {len(rows_above)}개 행 (1 ~ {start_row-1}), 아래쪽 {len(rows_below)}개 행 ({end_row+1} ~ {max_ws_row})")
    
    rows_with_merged_data_moved = process_merged_cells_outside_table(
        ws, start_row, end_row, start_col, end_col
    )
    
    # 4. 다른 행들 처리
    processed_rows = process_other_rows(
        ws, all_rows_to_process, start_col, end_col, rows_with_merged_data_moved
    )
    
    if processed_rows:
        logger.info(f"  - 다음 행 처리 완료: 총 {len(processed_rows)}개 행 (행 {min(processed_rows)} ~ {max(processed_rows)})")
    else:
        logger.info(f"  - 처리할 다음 행 없음")
    
    # 5. 첫 번째 시트의 모든 셀에 흰색 채우기 적용
    apply_white_fill_to_sheet(ws)
    
    # 6. 폰트 및 행 높이 설정
    apply_fonts_and_row_heights(ws, start_row, end_row)
    
    # 7. 모든 테두리 삭제
    remove_all_borders(ws)
    
    # 8. 테이블 테두리 및 스타일 적용
    apply_table_borders_and_styles(ws, start_row, end_row, start_col, end_col)
    
    # 9. 열 너비 설정
    set_table_column_widths(ws, start_col, end_col)
    
    # 10. 그림 배치
    place_image_at_center_top(ws, start_row, end_row, start_col, end_col)
    
    # 11. 모든 시트에서 텍스트 찾기/바꾸기
    replace_text_in_all_sheets(wb)
    
    # 파일 저장
    logger.info(f"\n파일 저장 중: {output_path}")
    wb.save(output_path)
    logger.info(f"완료! 저장된 파일: {output_path}")
    
    logger.info("=" * 60)
    logger.info("Excel 파일 통합 처리 완료!")
    logger.info("=" * 60)
    
    return output_path


def select_table_range(input_file: str) -> str:
    """
    Excel 파일을 열고 사용자가 테이블 범위를 선택하도록 합니다.
    
    Args:
        input_file: Excel 파일 경로
        
    Returns:
        선택된 범위 문자열 (예: "B34:AL38")
    """
    excel = None
    workbook = None
    
    try:
        # Excel 애플리케이션 시작
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = True
        excel.DisplayAlerts = False
        
        # 파일 열기
        input_path = Path(input_file)
        if not input_path.exists():
            raise FileNotFoundError(f"파일을 찾을 수 없습니다: {input_file}")
        
        workbook = excel.Workbooks.Open(str(input_path.absolute()))
        worksheet = workbook.ActiveSheet
        
        print("\n" + "=" * 60)
        print("테이블 범위 선택")
        print("=" * 60)
        print(f"파일: {input_path.name}")
        print(f"시트: {worksheet.Name}")
        print("\n엑셀에서 테이블 범위를 마우스로 드래그하여 선택하거나,")
        print("범위를 직접 입력하세요 (예: B34:AL38)")
        
        # Excel InputBox를 사용하여 범위 입력받기
        try:
            range_input = excel.InputBox(
                Prompt="테이블 범위를 선택하거나 입력하세요 (예: B34:AL38):\n범위를 마우스로 드래그하여 선택하세요.",
                Title="테이블 범위 선택",
                Type=8  # Range 타입
            )
            
            # InputBox가 False를 반환하면 취소된 것
            if range_input is False:
                print("\n범위 선택이 취소되었습니다.")
                return DEFAULT_RANGE
            
            if range_input:
                range_address = None
                
                # Range 객체인 경우 주소 가져오기
                if hasattr(range_input, 'Address'):
                    range_address = range_input.Address(False, False)
                else:
                    # 문자열인 경우
                    range_address = str(range_input).strip()
                
                # 범위 형식 검증
                if range_address and ':' in range_address:
                    print(f"\n선택된 범위: {range_address}")
                    return range_address
                else:
                    # 범위 형식이 아닌 경우 (예: 셀 값이 반환된 경우)
                    print(f"\n경고: '{range_address}'는 유효한 범위 형식이 아닙니다.")
                    print("기본값을 사용합니다.")
                    return DEFAULT_RANGE
            else:
                # None 또는 빈 값인 경우
                print("\n범위가 선택되지 않았습니다.")
                return DEFAULT_RANGE
                
        except Exception as e:
            logger.warning(f"InputBox 사용 실패: {e}")
            # InputBox 실패 시 선택된 범위 사용
            try:
                selection = excel.Selection
                if selection and hasattr(selection, 'Address'):
                    range_address = selection.Address(False, False)
                    if range_address and ':' in range_address:
                        print(f"\n선택된 범위: {range_address}")
                        return range_address
            except Exception as e2:
                logger.debug(f"선택된 범위 가져오기 실패: {e2}")
        
        # 기본값 반환
        print(f"\n기본값({DEFAULT_RANGE})을 사용합니다.")
        return DEFAULT_RANGE
        
    except Exception as e:
        logger.error(f"Excel 범위 선택 중 오류: {e}")
        print(f"\n오류 발생으로 기본값({DEFAULT_RANGE})을 사용합니다.")
        return DEFAULT_RANGE
        
    finally:
        # 리소스 정리
        try:
            if workbook:
                workbook.Close(SaveChanges=False)
            if excel:
                excel.Quit()
        except Exception as e:
            logger.debug(f"Excel 정리 중 오류 (무시): {e}")


if __name__ == "__main__":
    # 파일 탐색기를 띄워서 Excel 파일 선택
    root = tk.Tk()
    root.withdraw()  # 메인 윈도우 숨기기
    
    print("=" * 60)
    print("Excel 파일 선택")
    print("=" * 60)
    print("처리할 Excel 파일을 선택하세요.")
    
    input_file = filedialog.askopenfilename(
        title="Excel 파일 선택",
        filetypes=[
            ("Excel 파일", "*.xlsx *.xlsm *.xltx *.xltm *.xls"),
            ("모든 파일", "*.*")
        ],
        initialdir=r"C:\Users\u\Desktop\1219"
    )
    
    if not input_file:
        print("\n파일이 선택되지 않았습니다. 프로그램을 종료합니다.")
        exit(0)
    
    input_file = str(input_file)
    print(f"\n선택된 파일: {input_file}")
    
    try:
        # 사용자가 테이블 범위 선택
        selected_range = select_table_range(input_file)
        print(f"\n선택된 테이블 범위: {selected_range}")
        
        # 통합 처리: 1단계 + 2단계
        output_file = process_excel_file(input_file, range_str=selected_range)
        print(f"\n성공적으로 완료되었습니다!")
        print(f"최종 출력 파일: {output_file}")
    except Exception as e:
        logger.error(f"오류 발생: {e}")
        raise
